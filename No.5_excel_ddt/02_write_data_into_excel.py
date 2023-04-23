from openpyxl import load_workbook

# 1.返回一个文件对象
wb = load_workbook("testcase.xlsx")

# 2.指定获取表单,如果是wb.active则是默认获取第一个表单；想当于表单对象
ws = wb["login"]

# 3.cell()方法获取单元格对象
one_cell = ws.cell(4, 1)
one_cell.value = 3

# 4.修改数据后保存，文件对象.save(),文件需要是关闭状态，不然保存会报错
wb.save("testcase.xlsx")
