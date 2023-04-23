from openpyxl import load_workbook

wb = load_workbook("testcase.xlsx")

ws = wb["login"]

one_cell = ws.cell(1, 1)

# 读取指定单元格的数据
# 将数据读取出来，构造为嵌套字典的列表
# 存放用例字段
testcase_list = []
head_list = []  # 存放表头，就是excel第一行的数据

for row in range(1, ws.max_row+1):
    one_row_dict = {}
    for column in range(1, ws.max_column+1):
        one_cell_value = ws.cell(row, column).value   #获取每个单元格的值
        if row == 1:
            head_list.append(one_cell_value)#  获取excel文件中，当行数为1时，所有的列名，存放到head——list列表
        else:
            key = head_list[column-1]   # 遍历获取表头的值
            one_row_dict[key] = one_cell_value  # 赋值给key（key是表头），形成字典的键值对，添加到one_row_dict中
    if row != 1:
        testcase_list.append(one_row_dict)

print(testcase_list)


# title = []
# case = []
# for i in range(1, ws.max_row + 1):
#     parms_list = {}
#     for j in range(1, ws.max_column + 1):
#         parms = ws.cell(i, j).value
#         if i == 1:
#             title.append(parms)
#         else:
#             key = title[j - 1]
#             parms_list[key] = parms
#     if i != 1:
#         case.append(parms_list)
#
# print(case)

