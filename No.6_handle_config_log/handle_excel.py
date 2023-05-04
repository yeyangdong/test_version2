from openpyxl import load_workbook


class HandleExcel():
    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname

    def read_data(self):
        wb = load_workbook(self.filename)
        ws = wb[self.sheetname]

        testcase_list = []
        head_list = []  # 存放表头，就是excel第一行的数据

        for row in range(1, ws.max_row + 1):
            one_row_dict = {}
            for column in range(1, ws.max_column + 1):
                one_cell_value = ws.cell(row, column).value  # 获取每个单元格的值
                if row == 1:
                    head_list.append(one_cell_value)  # 获取excel文件中，当行数为1时，所有的列名，存放到head——list列表
                else:
                    key = head_list[column - 1]  # 遍历获取表头的值
                    one_row_dict[key] = one_cell_value  # 赋值给key（key是表头），形成字典的键值对
            if row != 1:
                testcase_list.append(one_row_dict)

        return testcase_list

    def write_data(self, row, column, data):
        """
        :param row:指定在某一行写
        :param column:指定在某一列写
        :param data:待写入的数据
        :return:
        1.如果将数据写入到excel中，不能读取操作公用一个Workbook对象
        """
        wb = load_workbook(self.filename)
        ws = wb[self.sheetname]

        # 第一种写入方式
        # one_cell = ws.cell(row,column)
        # one_cell.value = data

        # 第二种写入方式
        ws.cell(row, column, value=data)

        wb.save(self.filename)

if __name__ == '__main__':

    handle = HandleExcel("testcase.xlsx","register")
    handle.read_data()
    print(handle.read_data())
    handle.write_data(8,8,"charushuju")