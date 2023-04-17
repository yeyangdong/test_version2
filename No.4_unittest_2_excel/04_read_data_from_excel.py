from openpyxl import load_workbook

# excel对象--sheet表单对象--cell单元对象--行和列，值对象

# 1.返回一个文件对象
wb = load_workbook("testcase.xlsx")

# 2.获取表单,默认获取第一个表单，想当于表单对象
# 其中返回了最大和最小的行列数
ws = wb.active

# 3.cell()方法获取单元格，返回cell对象，行列的值，都从1开始计算
# 可以使用cell对象中.value方法获取单元格的值
one_cell = ws.cell(1, 1)

for row in range(1, ws.max_row+1):
    for column in range(1, ws.max_column+1):
        one_cell = ws.cell(row, column)
        print(f"单元格{row, column}中的值为{one_cell.value}")
